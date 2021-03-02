import requests
import secrets
import json
import sqlite3
from typing import Tuple
from openpyxl import load_workbook

workbook = load_workbook(filename="state_M2019_dl.xslx")
sheet = workbook.active

states = {}

def get_data():
    all_data = []
    for page in range(162):
        response = requests.get(f"https://api.data.gov/ed/collegescorecard/v1/schools.json?"
                                f"school.degrees_awarded.predominant=2,3&fields=id,school.name,school.city,"
                                f"2018.student.size,2017.student.size,"
                                f"2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line,"
                                f"2016.repayment.3_yr_repayment.overall&api_key={secrets.api_key}&page={page}")
        if response.status_code != 200:
            print("error getting data")

        page_of_data = response.json()
        page_of_school_data = page_of_data['results']
        all_data.extend(page_of_school_data)
    return all_data

def open_db(filename: str) -> Tuple[sqlite3.Connection, sqlite3.Cursor]:
        db_connection = sqlite3.connect(filename)#connect to existing DB or create new one
        cursor = db_connection.cursor()#get ready to read/write data
        return db_connection, cursor

def close_db(connection: sqlite3.Connection):
    connection.commit()  # make sure any changes get saved
    connection.close()

def setup_db(cursor:sqlite3.Cursor):
    cursor.execute('''CREATE TABLE IF NOT EXISTS college_data(
    school_id INTEGER PRIMARY KEY,
    school_name TEXT NOT NULL,
    school_city TEXT NOT NULL, 
    student_size INTEGER,
    three_year_earnings_over_poverty INTEGER,
    loan_repayment INTEGER
    );''')

def state_employment(cursor:sqlite3.Cursor):
    cursor.execute('''CREATE TABLE IF NOT EXISTS state_by_state_employment(
    employment_state TEXT NOT NULL,
    job_title TEXT NOT NULL,
    total_employment INTEGER,
    twenty_fifth_percentile_salary INTEGER,
    job_code''')

def add_college_table_data(all_data, cursor):
    for coll_data in all_data:
        cursor.execute("""
            INSERT INTO college_data(school_id, school_city, school_name, student_size, three_year_earnings_over_poverty,
             loan_repayment)
             VALUES (?,?,?,?,?,?);
            """, (coll_data['id'], coll_data['school.city'], coll_data['school.name'], coll_data['2018.student.size'],
                  coll_data['2017.earnings.3_yrs_after_completion.overall_count_over_poverty_line'],
                  coll_data['2016.repayment.3_yr_repayment.overall']))
   #for row in sheet.iter_rows():
     #   if sheet.columns[9].values="":
        #    cursor.execute("""
              #      INSERT INTO state_by_state_employment(employment_state, job_title, total_employment,
                #    twenty_fifth_percentile_salary, job_code)
                 #    VALUES (?,?,?,?,?,?);""", (
#                   row['state'], row['title'], row['tot_emp], row[twenty_five_percent'], row['code']


def main():
    all_data = get_data()
    conn, cursor = open_db("college_db.sqlite")
    setup_db(cursor)
    add_college_table_data(all_data, cursor)
    with open('school_data.txt', 'w') as outfile:
        json.dump(add_college_table_data(), outfile)
    print(add_college_table_data())
    print(json.dumps(states))
    close_db(conn)




if __name__ == '__main__':
    main()
